# Copyright (c) Hewlett Packard Enterprise, 2026. All rights reserved.
# SPDX-License-Identifier: Apache-2.0

"""End-to-end system integration tests for the NITA Webapp REST API.

These tests run against a *live* NITA webapp + MariaDB stack deployed on a
kind cluster.  They use real HTTP (``requests``) — no Django test-client or
mock layers.

Prerequisites (set up by the CI workflow before running pytest):
  • The webapp is reachable at ``NITA_BASE_URL`` (default http://localhost:8000).
  • The ``evpn_vxlan_erb_dc_1.3`` CampusType has been seeded via
    ``manage.py seed_evpn_fixture``.
  • ``EVPN_FIXTURE_DIR`` points at a directory with ``project.yaml``,
    ``dc1-hosts``, ``dc2-hosts``, and ``dc1_data.xlsx``.

Scope
-----
All tests stop short of triggering NOOB / build / test Jenkins jobs.
"""

import pathlib
import uuid

import pytest
import requests

from tests.integration.conftest import BASE_URL, FIXTURE_DIR


# ===========================================================================
# Section 1 — Project / CampusType (read-only checks on seeded data)
# ===========================================================================


@pytest.mark.integration
def test_auth_token_endpoint_returns_token(api_session):
    """The session fixture itself proved we can obtain a token; this test
    also verifies the token endpoint rejects invalid credentials.
    """
    resp = requests.post(
        f"{BASE_URL}/api/v1/auth/token/",
        json={"username": "__bad__", "password": "__bad__"},
        timeout=15,
    )
    assert resp.status_code == 400


@pytest.mark.integration
def test_unauthenticated_request_rejected():
    """Any API endpoint must refuse unauthenticated requests."""
    resp = requests.get(f"{BASE_URL}/api/v1/networks/", timeout=15)
    assert resp.status_code in (401, 403)


@pytest.mark.screenshot("/campustype/")
@pytest.mark.integration
def test_project_campus_type_visible_in_list(api_session, evpn_campus_type_id):
    """GET /api/v1/network-types/ must include the seeded evpn type."""
    resp = api_session.get(f"{BASE_URL}/api/v1/network-types/", timeout=30)
    assert resp.status_code == 200
    names = [r["name"] for r in resp.json()["results"]]
    assert "evpn_vxlan_erb_dc_1.3" in names


@pytest.mark.screenshot("/campustype/{evpn_campus_type_id}/")
@pytest.mark.integration
def test_project_campus_type_retrieve(api_session, evpn_campus_type_id):
    """GET /api/v1/network-types/{id}/ must return correct name/description."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/network-types/{evpn_campus_type_id}/", timeout=30
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["name"] == "evpn_vxlan_erb_dc_1.3"
    assert data["description"] == "EVPN VLXAN ERB Data Center"
    assert "roles" in data
    assert "resources" in data


@pytest.mark.screenshot("/campustype/{evpn_campus_type_id}/")
@pytest.mark.integration
def test_project_actions_listed_for_campus_type(api_session, evpn_campus_type_id):
    """GET /api/v1/actions/?campus_type_id=... must return the four actions
    defined in project.yaml.
    """
    resp = api_session.get(
        f"{BASE_URL}/api/v1/actions/?campus_type_id={evpn_campus_type_id}",
        timeout=30,
    )
    assert resp.status_code == 200
    results = resp.json()["results"]
    # project.yaml has exactly 4 actions: Build, Build base config,
    # Dump configuration, Test
    assert len(results) == 4
    action_names = {r["action_name"] for r in results}
    assert "Build" in action_names
    assert "Test" in action_names


@pytest.mark.screenshot("/campustype/{evpn_campus_type_id}/")
@pytest.mark.integration
def test_project_actions_reference_valid_categories(api_session, evpn_campus_type_id):
    """Every action for the evpn type must reference BUILD or TEST category."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/actions/?campus_type_id={evpn_campus_type_id}",
        timeout=30,
    )
    assert resp.status_code == 200
    for action in resp.json()["results"]:
        assert action["action_category"]["category_name"] in ("BUILD", "TEST")


@pytest.mark.integration
def test_action_categories_list(api_session):
    """GET /api/v1/action-categories/ must return BUILD and TEST entries."""
    resp = api_session.get(f"{BASE_URL}/api/v1/action-categories/", timeout=30)
    assert resp.status_code == 200
    names = {r["category_name"] for r in resp.json()["results"]}
    assert "BUILD" in names
    assert "TEST" in names


# ===========================================================================
# Section 2 — Network creation
# ===========================================================================


@pytest.mark.integration
def test_network_create_returns_201(api_session, evpn_campus_type_id):
    """POST /api/v1/networks/ with valid dc1-hosts must return 201 and
    the id/name of the new network.
    """
    host_content = (FIXTURE_DIR / "dc1-hosts").read_text()
    name = f"evpn-ci-create-{uuid.uuid4().hex[:8]}"
    payload = {
        "name": name,
        "status": "Initialized",
        "description": "Integration test network",
        "host_file": host_content,
        "campus_type": evpn_campus_type_id,
    }
    resp = api_session.post(f"{BASE_URL}/api/v1/networks/", json=payload, timeout=30)
    assert resp.status_code == 201
    data = resp.json()
    assert data["name"] == name
    assert "id" in data
    network_id = data["id"]

    # Cleanup
    api_session.delete(f"{BASE_URL}/api/v1/networks/{network_id}/", timeout=30)


@pytest.mark.integration
def test_network_create_with_dc2_hosts(api_session, evpn_campus_type_id):
    """The same CampusType can back a second network using dc2-hosts."""
    host_content = (FIXTURE_DIR / "dc2-hosts").read_text()
    name = f"evpn-ci-dc2-{uuid.uuid4().hex[:8]}"
    payload = {
        "name": name,
        "status": "Initialized",
        "description": "DC2 integration test",
        "host_file": host_content,
        "campus_type": evpn_campus_type_id,
    }
    resp = api_session.post(f"{BASE_URL}/api/v1/networks/", json=payload, timeout=30)
    assert resp.status_code == 201
    assert resp.json()["name"] == name
    api_session.delete(
        f"{BASE_URL}/api/v1/networks/{resp.json()['id']}/", timeout=30
    )


# ---------------------------------------------------------------------------
# 2.1  Host file load
# ---------------------------------------------------------------------------


@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_host_file_stored_and_retrievable(api_session, evpn_network):
    """GET /api/v1/networks/{id}/ must return the host_file exactly as POSTed."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/", timeout=30
    )
    assert resp.status_code == 200
    stored = resp.json()["host_file"]
    original = (FIXTURE_DIR / "dc1-hosts").read_text()
    assert stored.strip() == original.strip()


@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_host_file_contains_dc1_inventory_groups(api_session, evpn_network):
    """The host_file of a DC1 network must contain the expected Ansible
    inventory group names.
    """
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/", timeout=30
    )
    assert resp.status_code == 200
    host_file = resp.json()["host_file"]
    for group in ("spines", "leaves", "firewalls", "servers"):
        assert group in host_file, f"Group '{group}' missing from host_file"


@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_host_file_patch_with_dc2_hosts(api_session, evpn_network):
    """PATCH /api/v1/networks/{id}/ must accept a replacement host_file."""
    dc2_content = (FIXTURE_DIR / "dc2-hosts").read_text()
    resp = api_session.patch(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/",
        json={"host_file": dc2_content},
        timeout=30,
    )
    assert resp.status_code == 200
    # Verify the change was persisted
    get_resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/", timeout=30
    )
    assert get_resp.json()["host_file"].strip() == dc2_content.strip()


# ---------------------------------------------------------------------------
# 2.2  Configuration load (spreadsheet upload)
# ---------------------------------------------------------------------------


@pytest.mark.integration
def test_workbook_upload_dc1_xlsx_returns_200(api_session, evpn_network):
    """POST …/workbook/upload/ with dc1_data.xlsx must succeed (HTTP 200)."""
    with (FIXTURE_DIR / "dc1_data.xlsx").open("rb") as fh:
        resp = api_session.post(
            f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/workbook/upload/",
            files={"up_file": ("dc1_data.xlsx", fh, "application/octet-stream")},
            timeout=60,
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "success"


@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_upload_persists_base_sheet(api_session, evpn_network_with_workbook):
    """After upload, GET …/workbook/ must include a 'base' sheet."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/",
        timeout=30,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "success"
    sheet_names = [s["name"] for s in body["workbook"]]
    assert "base" in sheet_names


@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_upload_persists_all_expected_sheets(
    api_session, evpn_network_with_workbook
):
    """The workbook endpoint must expose all sheets from dc1_data.xlsx."""
    from openpyxl import load_workbook as openpyxl_load

    expected = set(openpyxl_load(FIXTURE_DIR / "dc1_data.xlsx").sheetnames)

    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/",
        timeout=30,
    )
    assert resp.status_code == 200
    returned = {s["name"] for s in resp.json()["workbook"]}
    assert expected == returned


@pytest.mark.integration
def test_workbook_upload_without_file_returns_400(api_session, evpn_network):
    """Uploading with no file must return HTTP 400."""
    resp = api_session.post(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/workbook/upload/",
        files={},
        timeout=30,
    )
    assert resp.status_code == 400


@pytest.mark.integration
def test_workbook_upload_idempotent(api_session, evpn_network):
    """Uploading the same workbook twice must not double the sheet count."""
    xlsx_path = FIXTURE_DIR / "dc1_data.xlsx"

    def _upload():
        with xlsx_path.open("rb") as fh:
            r = api_session.post(
                f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/workbook/upload/",
                files={"up_file": ("dc1_data.xlsx", fh, "application/octet-stream")},
                timeout=60,
            )
        assert r.status_code == 200
        return r

    _upload()
    get1 = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/workbook/", timeout=30
    )
    count1 = len(get1.json()["workbook"])

    _upload()
    get2 = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/workbook/", timeout=30
    )
    count2 = len(get2.json()["workbook"])

    assert count1 == count2


# ---------------------------------------------------------------------------
# 2.3  Export (workbook download)
# ---------------------------------------------------------------------------


@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_download_returns_xlsx_attachment(
    api_session, evpn_network_with_workbook
):
    """GET …/workbook/download/ must return an xlsx file as an attachment."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/download/",
        timeout=60,
    )
    assert resp.status_code == 200
    content_type = resp.headers.get("Content-Type", "")
    content_disposition = resp.headers.get("Content-Disposition", "")
    assert "spreadsheetml" in content_type or "octet-stream" in content_type
    assert "attachment" in content_disposition
    assert ".xlsx" in content_disposition
    # The body must be non-empty bytes
    assert len(resp.content) > 0


@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_download_content_is_valid_xlsx(
    api_session, evpn_network_with_workbook
):
    """The downloaded file must be parseable as a valid xlsx workbook."""
    import io

    from openpyxl import load_workbook as openpyxl_load

    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/download/",
        timeout=60,
    )
    assert resp.status_code == 200
    wb = openpyxl_load(io.BytesIO(resp.content))
    assert len(wb.sheetnames) > 0


@pytest.mark.integration
def test_workbook_download_without_auth_rejected():
    """GET …/workbook/download/ without auth must be rejected."""
    # We need any network ID — 1 is likely to exist in a seeded DB but we
    # just care that auth is checked, not that the network exists.
    resp = requests.get(
        f"{BASE_URL}/api/v1/networks/1/workbook/download/", timeout=15
    )
    assert resp.status_code in (401, 403)


# ---------------------------------------------------------------------------
# 2.4  Full CRUD
# ---------------------------------------------------------------------------


# LIST
@pytest.mark.screenshot("/campusnetwork/")
@pytest.mark.integration
def test_network_list_returns_paginated_envelope(api_session, evpn_network):
    """GET /api/v1/networks/ must return count + results."""
    resp = api_session.get(f"{BASE_URL}/api/v1/networks/", timeout=30)
    assert resp.status_code == 200
    body = resp.json()
    assert "count" in body
    assert "results" in body
    assert body["count"] >= 1


@pytest.mark.screenshot("/campusnetwork/")
@pytest.mark.integration
def test_network_list_includes_created_network(api_session, evpn_network):
    """The newly created network must appear in the list."""
    resp = api_session.get(f"{BASE_URL}/api/v1/networks/", timeout=30)
    assert resp.status_code == 200
    names = [r["name"] for r in resp.json()["results"]]
    assert evpn_network["name"] in names


# RETRIEVE
@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_network_retrieve_returns_campus_type_name(api_session, evpn_network):
    """GET /api/v1/networks/{id}/ must include campus_type_name."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/", timeout=30
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["campus_type_name"] == "evpn_vxlan_erb_dc_1.3"


@pytest.mark.integration
def test_network_retrieve_unknown_id_returns_404(api_session):
    """Retrieving a non-existent network ID must return 404."""
    resp = api_session.get(f"{BASE_URL}/api/v1/networks/999999/", timeout=30)
    assert resp.status_code == 404


# UPDATE (PUT)
@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_network_full_update(api_session, evpn_network):
    """PUT /api/v1/networks/{id}/ must replace all supplied fields."""
    payload = {
        "name": evpn_network["name"],
        "status": "Updated",
        "description": "PUT-replaced description",
        "host_file": evpn_network["host_file"],
        "campus_type": evpn_network["campus_type"],
    }
    resp = api_session.put(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/",
        json=payload,
        timeout=30,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "Updated"
    assert data["description"] == "PUT-replaced description"


# PARTIAL UPDATE (PATCH)
@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_network_partial_update_description(api_session, evpn_network):
    """PATCH must update only the description field."""
    new_desc = "Patched by integration test"
    resp = api_session.patch(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/",
        json={"description": new_desc},
        timeout=30,
    )
    assert resp.status_code == 200
    assert resp.json()["description"] == new_desc


@pytest.mark.screenshot("/campusnetwork/{evpn_network}/")
@pytest.mark.integration
def test_network_partial_update_status(api_session, evpn_network):
    """PATCH must update only the status field."""
    resp = api_session.patch(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/",
        json={"status": "Configured"},
        timeout=30,
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "Configured"


# DELETE
@pytest.mark.integration
def test_network_delete_returns_204(api_session, evpn_campus_type_id):
    """DELETE /api/v1/networks/{id}/ must return 204 and remove the record."""
    host_content = (FIXTURE_DIR / "dc1-hosts").read_text()
    name = f"evpn-ci-del-{uuid.uuid4().hex[:8]}"
    create_resp = api_session.post(
        f"{BASE_URL}/api/v1/networks/",
        json={
            "name": name,
            "status": "Initialized",
            "description": "Delete test",
            "host_file": host_content,
            "campus_type": evpn_campus_type_id,
        },
        timeout=30,
    )
    assert create_resp.status_code == 201
    network_id = create_resp.json()["id"]

    delete_resp = api_session.delete(
        f"{BASE_URL}/api/v1/networks/{network_id}/", timeout=30
    )
    assert delete_resp.status_code == 204

    # Verify it's gone
    get_resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{network_id}/", timeout=30
    )
    assert get_resp.status_code == 404


@pytest.mark.integration
def test_network_delete_without_auth_rejected(evpn_network):
    """DELETE without auth must be rejected."""
    resp = requests.delete(
        f"{BASE_URL}/api/v1/networks/{evpn_network['id']}/", timeout=15
    )
    assert resp.status_code in (401, 403)


# WORKBOOK RETRIEVE + SAVE + CLEAR
@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_retrieve_after_upload(api_session, evpn_network_with_workbook):
    """GET …/workbook/ must return sheets with status=success after upload."""
    resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/",
        timeout=30,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "success"
    assert isinstance(body["workbook"], list)
    assert len(body["workbook"]) > 0


@pytest.mark.screenshot("/campus_network/{evpn_network_with_workbook}/configuration_view/")
@pytest.mark.integration
def test_workbook_save_updates_data(api_session, evpn_network_with_workbook):
    """POST …/workbook/save/ must persist updated grid rows."""
    import json as json_lib

    new_data = [
        {
            "name": "base",
            "base": [
                {
                    "host": "group_vars/all.yaml",
                    "name": "OS_dir",
                    "value": "/updated/by/integration-test/",
                }
            ],
        }
    ]
    resp = api_session.post(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/save/",
        data=json_lib.dumps({"data": new_data}),
        headers={"Content-Type": "application/json"},
        timeout=30,
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "success"


@pytest.mark.integration
def test_workbook_clear_removes_data(api_session, evpn_network_with_workbook):
    """DELETE …/workbook/clear/ must remove all workbook data (HTTP 204)."""
    resp = api_session.delete(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/clear/",
        timeout=30,
    )
    assert resp.status_code == 204

    # After clearing, workbook retrieve should return empty or 404
    get_resp = api_session.get(
        f"{BASE_URL}/api/v1/networks/{evpn_network_with_workbook['id']}/workbook/",
        timeout=30,
    )
    if get_resp.status_code == 200:
        assert get_resp.json().get("workbook", []) == []


# ACTION HISTORY
@pytest.mark.integration
def test_action_history_list_returns_paginated_envelope(api_session):
    """GET /api/v1/action-history/ must return a paginated envelope."""
    resp = api_session.get(f"{BASE_URL}/api/v1/action-history/", timeout=30)
    assert resp.status_code == 200
    body = resp.json()
    assert "count" in body
    assert "results" in body


@pytest.mark.screenshot("/campus_network/{evpn_network}/action_history/")
@pytest.mark.integration
def test_action_history_filter_by_network(api_session, evpn_network):
    """GET /api/v1/action-history/?campus_network_id=... must return only
    history for that network (empty list when no jobs have been run).
    """
    resp = api_session.get(
        f"{BASE_URL}/api/v1/action-history/?campus_network_id={evpn_network['id']}",
        timeout=30,
    )
    assert resp.status_code == 200
    # No jobs have been triggered so there should be no history
    assert resp.json()["results"] == []
